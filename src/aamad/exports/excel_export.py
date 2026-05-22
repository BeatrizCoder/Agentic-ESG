import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import logging

logger = logging.getLogger(__name__)

# Color palette
PURPLE       = "7C3AED"
PURPLE_LIGHT = "EDE9FE"
GREEN        = "059669"
GREEN_LIGHT  = "D1FAE5"
RED          = "DC2626"
RED_LIGHT    = "FEE2E2"
YELLOW       = "D97706"
YELLOW_LIGHT = "FEF3C7"
BLUE         = "2563EB"
BLUE_LIGHT   = "DBEAFE"
GRAY_DARK    = "1F2937"
GRAY_MID     = "6B7280"
GRAY_LIGHT   = "F9FAFB"
WHITE        = "FFFFFF"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold=False, size=11, color="1F2937", italic=False, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)


def _border(style="thin", color="E5E7EB") -> Border:
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def generate_excel_report(
    tickets: list,
    metrics: dict,
    agent_metrics: list,
    cost_forecast: dict,
    resolution_time: list,
    period: str = "All time",
) -> BytesIO:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _create_dashboard_sheet(wb, metrics, cost_forecast, resolution_time, period)
    _create_tickets_sheet(wb, tickets)
    _create_agents_sheet(wb, agent_metrics)
    _create_cost_sheet(wb, cost_forecast, metrics)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _create_dashboard_sheet(wb, metrics, cost_forecast, resolution_time, period):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False

    widths = [2, 20, 15, 15, 15, 15, 15, 15, 2]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 8

    ws.merge_cells("B2:H2")
    title_cell = ws["B2"]
    title_cell.value = "Agentic Support Platform"
    title_cell.font = _font(bold=True, size=20, color=WHITE)
    title_cell.fill = _fill(PURPLE)
    title_cell.alignment = _align(h="left", v="center")

    ws.merge_cells("B3:H3")
    sub_cell = ws["B3"]
    sub_cell.value = (
        f"Analytics Report · {period} · "
        f"Generated {datetime.now().strftime('%d %b %Y %H:%M')}"
    )
    sub_cell.font = _font(size=11, color=GRAY_MID, italic=True)
    sub_cell.fill = _fill(PURPLE_LIGHT)
    sub_cell.alignment = _align(h="left", v="center")

    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 35
    ws.row_dimensions[7].height = 25
    ws.row_dimensions[8].height = 15

    kpis = [
        ("Total Tickets", metrics.get("total_runs", 0), BLUE, BLUE_LIGHT, ""),
        ("Resolved",      metrics.get("resolved", 0),   GREEN, GREEN_LIGHT, ""),
        ("Escalated",     metrics.get("escalated", 0),  RED,  RED_LIGHT,   ""),
        ("CSAT Score",    f"{metrics.get('csat_score', 0) or 0:.0f}%", PURPLE, PURPLE_LIGHT, ""),
    ]

    for (label, value, color, light, icon), col in zip(kpis, [2, 4, 6, 8]):
        col_letter = get_column_letter(col)
        next_col   = get_column_letter(col + 1)

        ws.merge_cells(f"{col_letter}6:{next_col}6")
        val_cell = ws[f"{col_letter}6"]
        val_cell.value = str(value)
        val_cell.font = _font(bold=True, size=18, color=color)
        val_cell.fill = _fill(light)
        val_cell.alignment = _align(h="center", v="center")
        val_cell.border = _border(color=color)

        ws.merge_cells(f"{col_letter}7:{next_col}7")
        lbl_cell = ws[f"{col_letter}7"]
        lbl_cell.value = label
        lbl_cell.font = _font(size=10, color=GRAY_MID)
        lbl_cell.fill = _fill(light)
        lbl_cell.alignment = _align(h="center", v="center")

    ws.row_dimensions[10].height = 25

    ws.merge_cells("B10:D10")
    hdr = ws["B10"]
    hdr.value = "Tickets by Category"
    hdr.font = _font(bold=True, size=12, color=WHITE)
    hdr.fill = _fill(GRAY_DARK)
    hdr.alignment = _align(h="center", v="center")

    for col, text in [(2, "Category"), (3, "Tickets"), (4, "% Total")]:
        cell = ws.cell(row=11, column=col, value=text)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(PURPLE)
        cell.alignment = _align(h="center")
        cell.border = _border()

    categories = metrics.get("by_category", [])
    total = metrics.get("total_runs", 1) or 1

    cat_colors = {
        "Order Issues":    BLUE_LIGHT,
        "Billing":         RED_LIGHT,
        "Account Access":  YELLOW_LIGHT,
        "Technical Issue": PURPLE_LIGHT,
        "General Support": GREEN_LIGHT,
    }

    for i, cat in enumerate(categories):
        row = 12 + i
        ws.row_dimensions[row].height = 20
        pct = round(cat.get("count", 0) / total * 100, 1)
        bg  = cat_colors.get(cat.get("category", ""), GRAY_LIGHT)

        for col, val in [
            (2, cat.get("category", "")),
            (3, cat.get("count", 0)),
            (4, f"{pct}%"),
        ]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(size=10)
            cell.fill = _fill(bg)
            cell.alignment = _align(h="center" if col > 2 else "left")
            cell.border = _border()

    if categories:
        pie = PieChart()
        pie.title = "Ticket Distribution"
        pie.style = 10
        pie.width = 12
        pie.height = 10

        data_ref = Reference(ws, min_col=3, max_col=3, min_row=11, max_row=11 + len(categories))
        cats_ref = Reference(ws, min_col=2, max_col=2, min_row=12, max_row=11 + len(categories))

        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)

        try:
            from openpyxl.chart.label import DataLabelList
            pie.dataLabels = DataLabelList()
            pie.dataLabels.showPercent = True
        except Exception:
            pass

        ws.add_chart(pie, "F10")

    start_row = 12 + len(categories) + 2

    ws.merge_cells(f"B{start_row}:D{start_row}")
    rt_hdr = ws[f"B{start_row}"]
    rt_hdr.value = "Avg Resolution Time by Category"
    rt_hdr.font = _font(bold=True, size=12, color=WHITE)
    rt_hdr.fill = _fill(GRAY_DARK)
    rt_hdr.alignment = _align(h="center", v="center")
    ws.row_dimensions[start_row].height = 25

    for col, text in [(2, "Category"), (3, "Avg Time (s)"), (4, "Speed")]:
        cell = ws.cell(row=start_row + 1, column=col, value=text)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(PURPLE)
        cell.alignment = _align(h="center")
        cell.border = _border()

    for i, rt in enumerate(resolution_time or []):
        row = start_row + 2 + i
        ws.row_dimensions[row].height = 20

        avg = rt.get("avg_time_sec", 0)
        speed = "Fast" if avg < 8 else "Medium" if avg < 12 else "Slow"
        speed_color = GREEN_LIGHT if avg < 8 else YELLOW_LIGHT if avg < 12 else RED_LIGHT

        for col, val in [(2, rt.get("category", "")), (3, avg), (4, speed)]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = _font(size=10)
            cell.fill = _fill(speed_color)
            cell.alignment = _align(h="center" if col > 2 else "left")
            cell.border = _border()

    return ws


def _create_tickets_sheet(wb, tickets):
    ws = wb.create_sheet("🎫 Tickets")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    columns = [
        ("Reference ID", 18),
        ("Date", 14),
        ("Category", 16),
        ("Sentiment", 12),
        ("Urgency", 10),
        ("Status", 16),
        ("Routing", 14),
        ("Response Time (s)", 16),
        ("CSAT", 8),
        ("Quality Grade", 12),
    ]

    for i, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 30
    for i, (name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.font = _font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(PURPLE)
        cell.alignment = _align(h="center", v="center")
        cell.border = _border()

    status_styles = {
        "completed":            (GREEN, GREEN_LIGHT),
        "pending_human_review": (RED,   RED_LIGHT),
        "awaiting_customer_info": (YELLOW, YELLOW_LIGHT),
        "approved":             (GREEN, GREEN_LIGHT),
        "rejected":             (RED,   RED_LIGHT),
    }

    for i, ticket in enumerate(tickets, 2):
        ws.row_dimensions[i].height = 18

        if isinstance(ticket, dict):
            t = ticket
        else:
            t = ticket.model_dump() if hasattr(ticket, "model_dump") else ticket.__dict__

        created = t.get("created_at", "")
        if isinstance(created, str) and created:
            try:
                dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                created = dt.strftime("%d/%m/%Y %H:%M")
            except Exception:
                pass

        status = t.get("status", "")
        _, bg = status_styles.get(status, (GRAY_MID, GRAY_LIGHT))

        quality = t.get("quality_evaluation") or {}
        if isinstance(quality, str):
            try:
                quality = json.loads(quality)
            except Exception:
                quality = {}
        grade = quality.get("grade", "—") if isinstance(quality, dict) else "—"

        feedback = t.get("feedback") or {}
        if isinstance(feedback, str):
            try:
                feedback = json.loads(feedback)
            except Exception:
                feedback = {}
        if isinstance(feedback, dict):
            csat = "Helpful" if feedback.get("helpful") is True else "Not Helpful" if feedback.get("helpful") is False else "—"
        else:
            csat = str(feedback) if feedback else "—"

        row_data = [
            t.get("reference_id", ""),
            created,
            t.get("category", ""),
            t.get("sentiment", ""),
            t.get("urgency", ""),
            status.replace("_", " ").title(),
            t.get("routing_action", "") or "",
            round((t.get("execution_time_ms") or 0) / 1000, 1),
            csat,
            grade,
        ]

        for j, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=value)
            cell.font = _font(size=10)
            cell.fill = _fill(bg if j == 6 else WHITE if i % 2 == 0 else GRAY_LIGHT)
            cell.alignment = _align(h="center" if j != 1 else "left")
            cell.border = _border()

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(tickets) + 1}"
    return ws


def _create_agents_sheet(wb, agent_metrics):
    ws = wb.create_sheet("🤖 Agent Performance")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "CrewAI Agent Performance Report"
    title.font = _font(bold=True, size=14, color=WHITE)
    title.fill = _fill(PURPLE)
    title.alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 35

    widths = [25, 10, 15, 15, 15, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = [
        "Agent", "Calls", "Avg Latency (ms)",
        "Avg Input Tokens", "Avg Output Tokens",
        "Cost/Call (USD)", "Execution Mode",
    ]

    ws.row_dimensions[2].height = 25
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(GRAY_DARK)
        cell.alignment = _align(h="center", v="center")
        cell.border = _border()

    for i, agent in enumerate(agent_metrics or [], 3):
        ws.row_dimensions[i].height = 20
        bg = PURPLE_LIGHT if i % 2 == 0 else WHITE

        mode = agent.get("execution_mode", "llm")
        mode_label = "LLM" if mode == "llm" else "Rules" if mode == "deterministic" else "API"

        row_data = [
            agent.get("step_name", agent.get("agent_name", "")),
            agent.get("calls", 0),
            round(agent.get("avg_latency_ms", 0), 0),
            round(agent.get("avg_input_tokens", 0), 0),
            round(agent.get("avg_output_tokens", 0), 0),
            agent.get("avg_cost_usd", 0),
            mode_label,
        ]

        for j, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = _font(size=10)
            cell.fill = _fill(bg)
            cell.alignment = _align(h="left" if j == 1 else "center")
            cell.border = _border()

    if agent_metrics:
        bar = BarChart()
        bar.type = "col"
        bar.title = "Cost per Agent (USD)"
        bar.style = 10
        bar.width = 20
        bar.height = 12

        last_row = 2 + len(agent_metrics)
        data = Reference(ws, min_col=6, max_col=6, min_row=2, max_row=last_row)
        cats = Reference(ws, min_col=1, max_col=1, min_row=3, max_row=last_row)

        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        try:
            bar.series[0].graphicalProperties.solidFill = PURPLE
        except Exception:
            pass

        ws.add_chart(bar, f"A{last_row + 3}")

    return ws


def _create_cost_sheet(wb, cost_forecast, metrics):
    ws = wb.create_sheet("💰 Cost Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "Cost Forecasting & Analysis"
    title.font = _font(bold=True, size=14, color=WHITE)
    title.fill = _fill(GREEN)
    title.alignment = _align(h="center", v="center")
    ws.row_dimensions[1].height = 35

    for i, w in enumerate([25, 15, 15, 15, 15, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A3:F3")
    proj_hdr = ws["A3"]
    proj_hdr.value = "Cost Projections"
    proj_hdr.font = _font(bold=True, size=12, color=WHITE)
    proj_hdr.fill = _fill(GRAY_DARK)
    proj_hdr.alignment = _align(h="center", v="center")
    ws.row_dimensions[3].height = 25

    projections = cost_forecast.get("projections", {})
    avg_daily   = cost_forecast.get("avg_daily_tickets", 0)

    proj_data = [
        ("Daily",   projections.get("daily_usd", 0),   round(avg_daily, 1)),
        ("Weekly",  projections.get("weekly_usd", 0),  round(avg_daily * 7, 1)),
        ("Monthly", projections.get("monthly_usd", 0), round(avg_daily * 30, 1)),
        ("Yearly",  projections.get("yearly_usd", 0),  round(avg_daily * 365, 1)),
    ]

    for i, text in enumerate(["Period", "Projected Cost (USD)", "Est. Tickets"], 1):
        cell = ws.cell(row=4, column=i, value=text)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(GREEN)
        cell.alignment = _align(h="center")
        cell.border = _border()
    ws.row_dimensions[4].height = 22

    for i, (period, cost, tickets) in enumerate(proj_data, 5):
        ws.row_dimensions[i].height = 20
        bg = GREEN_LIGHT if i % 2 == 0 else WHITE

        for j, val in enumerate([period, f"${cost:.4f}", tickets], 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = _font(size=11, bold=(j == 2), color=GREEN if j == 2 else GRAY_DARK)
            cell.fill = _fill(bg)
            cell.alignment = _align(h="left" if j == 1 else "center")
            cell.border = _border()

    ws.merge_cells("A10:F10")
    breakdown_hdr = ws["A10"]
    breakdown_hdr.value = "Cost Breakdown by Agent"
    breakdown_hdr.font = _font(bold=True, size=12, color=WHITE)
    breakdown_hdr.fill = _fill(GRAY_DARK)
    breakdown_hdr.alignment = _align(h="center", v="center")
    ws.row_dimensions[10].height = 25

    agent_costs = [
        ("Classification Agent",    0.0002, "6.5%"),
        ("Sentiment Agent",         0.0003, "9.7%"),
        ("Knowledge Agent",         0.0000, "0% (local)"),
        ("Response Agent",          0.0012, "38.7%"),
        ("Quality Agent (Sonnet)",  0.0012, "38.7%"),
        ("Summary Agent",           0.0002, "6.5%"),
    ]

    for j, text in enumerate(["Agent", "Cost/Ticket (USD)", "% of Total"], 1):
        cell = ws.cell(row=11, column=j, value=text)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(PURPLE)
        cell.alignment = _align(h="center")
        cell.border = _border()
    ws.row_dimensions[11].height = 22

    for i, (agent, cost, pct) in enumerate(agent_costs, 12):
        ws.row_dimensions[i].height = 20
        bg = PURPLE_LIGHT if i % 2 == 0 else WHITE

        for j, val in enumerate([agent, f"${cost:.4f}", pct], 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = _font(size=10)
            cell.fill = _fill(bg)
            cell.alignment = _align(h="left" if j == 1 else "center")
            cell.border = _border()

    total_row = 12 + len(agent_costs)
    ws.row_dimensions[total_row].height = 22
    for j, val in enumerate(["TOTAL", "$0.0031", "100%"], 1):
        cell = ws.cell(row=total_row, column=j, value=val)
        cell.font = _font(bold=True, size=11, color=WHITE)
        cell.fill = _fill(GRAY_DARK)
        cell.alignment = _align(h="left" if j == 1 else "center")
        cell.border = _border()

    return ws
