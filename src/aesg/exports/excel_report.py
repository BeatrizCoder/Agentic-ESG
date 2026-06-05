"""Agentic ESG — Excel report generator (openpyxl)."""

import logging
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Palette ───────────────────────────────────────────────────────────────────
_GREEN_HEX   = "3A6B35"
_HDR_FILL    = PatternFill("solid", fgColor=_GREEN_HEX)
_HDR_FONT    = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
_BOLD_GRN    = Font(name="Calibri", color=_GREEN_HEX, bold=True, size=11)
_BOLD        = Font(name="Calibri", bold=True, size=9)
_NORMAL      = Font(name="Calibri", size=9)
_ALT_FILL    = PatternFill("solid", fgColor="F7F5F0")
_NASA_FILL   = PatternFill("solid", fgColor="D6EAF8")
_PROJ_FILL   = PatternFill("solid", fgColor="FAE5D3")
_HDR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)
_WRAP_TOP    = Alignment(horizontal="left", vertical="top", wrap_text=True)
_CENTER      = Alignment(horizontal="center", vertical="center")


def _border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _set_header(ws, row: int, headers: list[str]) -> None:
    for col, val in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = _HDR_ALIGN
        c.border    = _border()
    ws.row_dimensions[row].height = 18


def _autofit(ws, min_w: int = 10, max_w: int = 60) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
        letter  = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(max_len + 2, min_w), max_w)


def _analysis_period_rows(analysis: dict) -> list[tuple[str, str]]:
    pm = analysis.get("pipeline_metadata", {}) or {}
    if analysis.get("period_1") and analysis.get("period_2"):
        p1 = analysis["period_1"]
        p2 = analysis["period_2"]
        return [
            ("Comparison Period 1", f"{p1.get('start_year','—')}–{p1.get('end_year','—')}"),
            ("Comparison Period 2", f"{p2.get('start_year','—')}–{p2.get('end_year','—')}"),
        ]
    if pm.get("nasa_start_year") is not None or pm.get("nasa_end_year") is not None:
        return [
            ("Analysis Period Start", pm.get("nasa_start_year", "—")),
            ("Analysis Period End",   pm.get("nasa_end_year",   "—")),
        ]
    return []


def _inv_status(score: int) -> str:
    if score <= 40:  return "Investment Approved"
    if score <= 70:  return "Investment Conditioned"
    if score <= 85:  return "Investment Restricted"
    return "Investment Suspended"


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb: Workbook, analysis: dict) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = True

    # Branding header
    ws["A1"] = "Agentic ESG · ESG Climate Risk Intelligence"
    ws["A1"].font = _BOLD_GRN
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 22

    ws["A2"] = f"Generated {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC"
    ws["A2"].font = Font(name="Calibri", italic=True, color="888888", size=8)
    ws.merge_cells("A2:C2")

    score = analysis.get("risk_score", 0)
    conf  = analysis.get("confidence_score", 0)

    base_rows = [
        ("Region",            analysis.get("region_label", "—")),
        ("Latitude",          analysis.get("latitude", 0)),
        ("Longitude",         analysis.get("longitude", 0)),
        ("Analysis ID",       analysis.get("analysis_id", "—")),
        ("Date",              analysis.get("created_at", "")[:10]),
        ("Sector",            analysis.get("sector", "—")),
    ]
    rows = base_rows + _analysis_period_rows(analysis) + [
        ("Risk Score",        f"{score}/100"),
        ("Risk Level",        analysis.get("risk_level", "—").upper()),
        ("Risk Badge",        analysis.get("risk_badge_label", "—")),
        ("Investment Status", _inv_status(score)),
        ("Confidence Score",  f"{conf}%"),
        ("Pipeline Duration", f"{analysis.get('pipeline_duration_sec', 0):.1f}s"),
    ]

    for i, (key, val) in enumerate(rows):
        row_num = i + 4
        kc = ws.cell(row=row_num, column=1, value=key)
        vc = ws.cell(row=row_num, column=2, value=val)
        kc.font = _BOLD
        vc.font = _NORMAL
        kc.border = vc.border = _border()
        if i % 2 == 0:
            kc.fill = vc.fill = _ALT_FILL

    # Key metrics sub-table
    km = analysis.get("key_metrics", {})
    if km:
        start = len(rows) + 5
        lbl = ws.cell(row=start, column=1, value="KEY METRICS")
        lbl.font = Font(name="Calibri", color=_GREEN_HEX, bold=True, size=10)
        ws.merge_cells(f"A{start}:B{start}")

        _set_header(ws, start + 1, ["Metric", "Value"])
        metrics = [
            ("Temperature Change",  km.get("temp_change_label", "—")),
            ("Precipitation Change", km.get("precip_change_label", "—")),
            ("Compliance Exposure",  km.get("compliance_exposure_label", "—")),
            ("Hottest Year",         str(km.get("hottest_year", "—"))),
            ("Driest Year",          str(km.get("driest_year", "—"))),
        ]
        for j, (k, v) in enumerate(metrics):
            r = start + 2 + j
            kc = ws.cell(row=r, column=1, value=k)
            vc = ws.cell(row=r, column=2, value=v)
            kc.font = _BOLD
            vc.font = _NORMAL
            kc.border = vc.border = _border()
            if j % 2 == 0:
                kc.fill = vc.fill = _ALT_FILL

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 16


def _sheet_annual(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("Annual Climate Data")

    headers = ["Year", "Temp Mean (°C)", "Precipitation (mm)", "Solar (kWh/m²/day)", "Source"]
    _set_header(ws, 1, headers)

    for i, r in enumerate(analysis.get("annual_records", []), 2):
        source  = r.get("source", "nasa")
        temp    = r.get("temp_mean_celsius") or r.get("temp_mean_c")
        precip  = r.get("precip_total_mm")
        solar   = r.get("solar_mean_kwh_m2")

        values = [r.get("year"), temp, precip, solar, source]
        fill   = _NASA_FILL if source == "nasa" else _PROJ_FILL

        for col, val in enumerate(values, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font      = _NORMAL
            c.fill      = fill
            c.border    = _border()
            c.alignment = _CENTER
            if isinstance(val, float):
                c.number_format = "0.00"

    # Legend
    legend_row = len(analysis.get("annual_records", [])) + 3
    ws.cell(row=legend_row, column=1, value="■ NASA POWER (observed)").fill = _NASA_FILL
    ws.cell(row=legend_row, column=1).font = _NORMAL
    ws.cell(row=legend_row + 1, column=1, value="■ IPCC Projection (OpenMeteo)").fill = _PROJ_FILL
    ws.cell(row=legend_row + 1, column=1).font = _NORMAL

    _autofit(ws, min_w=12, max_w=30)


def _sheet_compliance(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("ESG Compliance")

    _set_header(ws, 1, ["Framework", "Exposure Level", "Summary", "Articles / Scenarios"])

    comp  = analysis.get("compliance_mapping", {})
    items = [
        ("CSRD",        comp.get("csrd_exposure","—"),           comp.get("csrd_summary",""),        comp.get("csrd_articles",[])),
        ("ISSB S2",     comp.get("issb_s2_exposure","—"),        comp.get("issb_s2_summary",""),     comp.get("issb_s2_scenarios",[])),
        ("EU Taxonomy", comp.get("eu_taxonomy_alignment","—"),   comp.get("eu_taxonomy_summary",""), comp.get("eu_taxonomy_criteria",[])),
    ]

    for i, (fw, exp, summ, arts) in enumerate(items, 2):
        arts_str = ", ".join((arts or [])[:6])
        row_data = [fw, str(exp).upper(), summ, arts_str]
        fill = _ALT_FILL if i % 2 == 0 else None

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font      = _BOLD if col == 1 else _NORMAL
            c.border    = _border()
            c.alignment = _WRAP_TOP
            if fill:
                c.fill = fill

        ws.row_dimensions[i].height = 48

    _autofit(ws, min_w=14, max_w=80)
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18


def _sheet_recs(wb: Workbook, analysis: dict) -> None:
    ws = wb.create_sheet("Recommendations")

    _set_header(ws, 1, ["Rank", "Framework", "Action", "Article", "Timeline", "Priority"])

    for i, r in enumerate(analysis.get("recommendations", []), 2):
        row_data = [
            r.get("rank", i - 1),
            r.get("framework", ""),
            r.get("action", ""),
            r.get("article", ""),
            r.get("timeline", "").replace("_", " "),
            r.get("priority", ""),
        ]
        fill = _ALT_FILL if i % 2 == 0 else None

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font      = _BOLD if col in (1, 2) else _NORMAL
            c.border    = _border()
            c.alignment = _WRAP_TOP
            if fill:
                c.fill = fill

        ws.row_dimensions[i].height = 36

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 16
    _autofit(ws, min_w=10, max_w=70)


# ── Public entry points ───────────────────────────────────────────────────────

def generate_batch_excel(batch_data: dict) -> BytesIO:
    """Generate a single-sheet Excel summary for batch analysis results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = True

    ws["A1"] = "Agentic ESG · Batch ESG Analysis"
    ws["A1"].font = _BOLD_GRN
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 22

    completed = batch_data.get("completed", 0)
    total     = batch_data.get("total", 0)
    ws["A2"] = (
        f"Generated {datetime.utcnow().strftime('%d %b %Y %H:%M')} UTC  ·  "
        f"{completed}/{total} regions completed"
    )
    ws["A2"].font = Font(name="Calibri", italic=True, color="888888", size=8)
    ws.merge_cells("A2:F2")

    _set_header(ws, 4, ["Region", "Risk Score", "Risk Level", "Investment Status", "Confidence", "Scenario"])

    results = sorted(
        batch_data.get("results", []),
        key=lambda r: r.get("risk_score", 0),
        reverse=True,
    )

    _ERR_FILL = PatternFill("solid", fgColor="FDEDEC")
    for i, r in enumerate(results, 5):
        is_err = r.get("status") == "error"
        row_data = [
            r.get("region", ""),
            "Error" if is_err else r.get("risk_score", 0),
            "—"     if is_err else r.get("risk_level", "").upper(),
            (r.get("error") or "")[:60] if is_err else r.get("investment_status", "—"),
            "—"     if is_err else f"{r.get('confidence_score', 0)}%",
            r.get("scenario", "—"),
        ]
        fill = _ERR_FILL if is_err else (_ALT_FILL if i % 2 == 0 else None)
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font      = _BOLD if col == 1 else _NORMAL
            c.border    = _border()
            c.alignment = _CENTER if col > 1 else _WRAP_TOP
            if fill:
                c.fill = fill

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Batch Excel generated: %d rows", len(results))
    return buf


def generate_excel(analysis: dict) -> BytesIO:
    wb = Workbook()
    _sheet_summary(wb, analysis)
    _sheet_annual(wb, analysis)
    _sheet_compliance(wb, analysis)
    _sheet_recs(wb, analysis)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info("Excel generated for analysis %s", analysis.get("analysis_id"))
    return buf
