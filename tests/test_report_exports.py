from io import BytesIO

from openpyxl import load_workbook

from src.aesg.exports.excel_report import generate_excel
from src.aesg.exports.pdf_report import _analysis_period_table, _styles, generate_pdf


def _find_summary_value(ws, label: str):
    for row in ws.iter_rows(min_row=4, max_row=20, min_col=1, max_col=2, values_only=True):
        if row[0] == label:
            return row[1]
    raise AssertionError(f"Label {label!r} not found in summary sheet")


def test_excel_export_includes_analysis_period():
    analysis = {
        "analysis_id": "CS-100",
        "region_label": "Test Region",
        "latitude": 0.0,
        "longitude": 0.0,
        "risk_score": 45,
        "risk_level": "medium",
        "risk_badge_label": "Conditioned",
        "confidence_score": 88,
        "pipeline_duration_sec": 10.2,
        "sector": "General",
        "created_at": "2026-06-02T00:00:00Z",
        "pipeline_metadata": {
            "nasa_start_year": 2010,
            "nasa_end_year": 2020,
        },
        "key_metrics": {},
        "annual_records": [],
        "compliance_mapping": {},
        "recommendations": [],
        "executive_summary": "",
    }

    buffer = generate_excel(analysis)
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Summary"]

    assert _find_summary_value(ws, "Analysis Period Start") == 2010
    assert _find_summary_value(ws, "Analysis Period End") == 2020


def test_pdf_export_includes_analysis_period():
    analysis = {
        "analysis_id": "CS-101",
        "region_label": "Test Region",
        "latitude": 0.0,
        "longitude": 0.0,
        "risk_score": 55,
        "risk_level": "medium",
        "risk_badge_label": "Conditioned",
        "confidence_score": 75,
        "pipeline_duration_sec": 8.5,
        "sector": "General",
        "created_at": "2026-06-02T00:00:00Z",
        "pipeline_metadata": {
            "nasa_start_year": 2015,
            "nasa_end_year": 2025,
        },
        "key_metrics": {},
        "annual_records": [],
        "compliance_mapping": {},
        "recommendations": [],
        "executive_summary": "",
    }

    styles = _styles()
    table = _analysis_period_table(analysis, width=450, styles=styles)
    assert table is not None
    assert table._cellvalues[0][0].text == "Analysis Period"
    assert "2015" in table._cellvalues[0][1].text
    assert "2025" in table._cellvalues[0][1].text

    buffer = generate_pdf(analysis)
    content = buffer.getvalue()
    assert len(content) > 1000


def test_excel_export_supports_comparison_periods():
    analysis = {
        "analysis_id": "CS-102",
        "region_label": "Test Region",
        "latitude": 0.0,
        "longitude": 0.0,
        "risk_score": 65,
        "risk_level": "high",
        "risk_badge_label": "Restricted",
        "confidence_score": 62,
        "pipeline_duration_sec": 12.0,
        "sector": "General",
        "created_at": "2026-06-02T00:00:00Z",
        "period_1": {"start_year": 2000, "end_year": 2010},
        "period_2": {"start_year": 2011, "end_year": 2021},
        "key_metrics": {},
        "annual_records": [],
        "compliance_mapping": {},
        "recommendations": [],
        "executive_summary": "",
    }

    buffer = generate_excel(analysis)
    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Summary"]

    assert _find_summary_value(ws, "Comparison Period 1") == "2000–2010"
    assert _find_summary_value(ws, "Comparison Period 2") == "2011–2021"
